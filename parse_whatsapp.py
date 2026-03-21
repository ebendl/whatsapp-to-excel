import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Regex for format 1: [2026-01-05, 09:35:42] Sender: Message (handles both / and - in dates)
RE_FMT1 = re.compile(r'^\[(\d{4}[-/]\d{2}[-/]\d{2}, \d{2}:\d{2}:\d{2})\] ([^:]+): (.*)')
RE_FMT1_SYS = re.compile(r'^\[(\d{4}[-/]\d{2}[-/]\d{2}, \d{2}:\d{2}:\d{2})\] ([^:]+)$')

# Regex for format 2: 2026/02/20, 20:37 - Sender: Message
RE_FMT2 = re.compile(r'^(\d{4}/\d{2}/\d{2}, \d{2}:\d{2}) - ([^:]+): (.*)')
RE_FMT2_SYS = re.compile(r'^(\d{4}/\d{2}/\d{2}, \d{2}:\d{2}) - (.+)$')

# Attachment patterns
RE_ATTACH_FMT1 = re.compile(r'<attached: ([^>]+)>')
RE_ATTACH_FMT2 = re.compile(r'(.+?) \(file attached\)')


def detect_format(lines):
    for line in lines[:20]:
        if RE_FMT1.match(line) or RE_FMT1_SYS.match(line):
            return 1
        if RE_FMT2.match(line) or RE_FMT2_SYS.match(line):
            return 2
    return 2


def is_new_message(line, fmt):
    if fmt == 1:
        return bool(RE_FMT1.match(line) or RE_FMT1_SYS.match(line))
    else:
        return bool(RE_FMT2.match(line) or RE_FMT2_SYS.match(line))


def parse_chat(filepath):
    with open(filepath, encoding='utf-8') as f:
        lines = f.readlines()

    fmt = detect_format(lines)
    messages = []
    current = None

    for raw_line in lines:
        line = raw_line.rstrip('\n')
        # Strip zero-width characters and LTR marks from the start of the line
        line = re.sub(r'^[\u200e\u202a\ufeff\u200b\u200c\u200d\u202c]+', '', line)

        if is_new_message(line, fmt):
            if current:
                messages.append(current)

            if fmt == 1:
                m = RE_FMT1.match(line) or RE_FMT1.match(line.lstrip('\u200e'))
                if not m:
                    # system message
                    ms = RE_FMT1_SYS.match(line)
                    if ms:
                        current = {'datetime': ms.group(1), 'sender': '', 'text': ms.group(2), 'attachments': []}
                    else:
                        current = None
                    continue
                current = {
                    'datetime': m.group(1),
                    'sender': m.group(2).strip(),
                    'text': m.group(3),
                    'attachments': []
                }
            else:
                m = RE_FMT2.match(line)
                if not m:
                    ms = RE_FMT2_SYS.match(line)
                    if ms:
                        current = {'datetime': ms.group(1), 'sender': '', 'text': ms.group(2), 'attachments': []}
                    else:
                        current = None
                    continue
                current = {
                    'datetime': m.group(1),
                    'sender': m.group(2).strip(),
                    'text': m.group(3),
                    'attachments': []
                }
        else:
            # continuation line
            if current is not None:
                current['text'] += '\n' + line

    if current:
        messages.append(current)

    # Post-process: extract attachments
    for msg in messages:
        text = msg['text']
        attachments = []

        if fmt == 1:
            # Extract <attached: filename> references
            found = RE_ATTACH_FMT1.findall(text)
            attachments.extend(found)
            # Clean attachment tags from text
            text = RE_ATTACH_FMT1.sub('', text).strip()
        else:
            # Check for "filename (file attached)" in the text
            # We'll split by lines to handle each line potentially being an attachment
            lines = text.split('\n')
            new_lines = []
            i = 0
            while i < len(lines):
                line = lines[i].strip()
                m = RE_ATTACH_FMT2.search(line)
                if m:
                    filename = m.group(1).strip()
                    attachments.append(filename)
                    
                    # Remove the "(file attached)" part and the filename from this line
                    line_without_attachment = line.replace(m.group(0), "").strip()
                    if line_without_attachment:
                        new_lines.append(line_without_attachment)
                    
                    # Check if next line is exactly the filename (common in document attachments)
                    if i + 1 < len(lines) and lines[i+1].strip() == filename:
                        i += 1 # skip next line
                else:
                    new_lines.append(lines[i])
                i += 1
            text = '\n'.join(new_lines).strip()

        msg['text'] = text
        msg['attachments'] = attachments

    return messages


def create_xlsx(messages, output_path, sheet_name='Chat'):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    # Header style
    header_font = Font(name='Arial', bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', start_color='2F5496')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    headers = ['Date & Time', 'Sender', 'Message', 'Attachments']
    col_widths = [22, 28, 80, 50]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 20

    # Data style
    data_font = Font(name='Arial', size=10)
    data_align_wrap = Alignment(vertical='top', wrap_text=True)
    data_align_nowrap = Alignment(vertical='top', wrap_text=False)

    for row_idx, msg in enumerate(messages, 2):
        dt = msg['datetime']
        sender = msg['sender']
        text = msg['text']
        attachments = ', '.join(msg['attachments']) if msg['attachments'] else ''

        ws.cell(row=row_idx, column=1, value=dt).font = data_font
        ws.cell(row=row_idx, column=1).alignment = data_align_nowrap

        ws.cell(row=row_idx, column=2, value=sender).font = data_font
        ws.cell(row=row_idx, column=2).alignment = data_align_nowrap

        ws.cell(row=row_idx, column=3, value=text).font = data_font
        ws.cell(row=row_idx, column=3).alignment = data_align_wrap

        ws.cell(row=row_idx, column=4, value=attachments).font = data_font
        ws.cell(row=row_idx, column=4).alignment = data_align_wrap

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(output_path)
    print(f'Saved: {output_path} ({len(messages)} messages)')

import sys
import argparse

def main():
    parser = argparse.ArgumentParser(description='Convert WhatsApp chat directory to Excel.')
    parser.add_argument('directory', help='Path to the unzipped WhatsApp chat directory')
    args = parser.parse_args()

    dir_path = os.path.abspath(args.directory)
    if not os.path.isdir(dir_path):
        print(f"Error: {dir_path} is not a directory.")
        sys.exit(1)

    # Find .txt file in the directory
    txt_files = [f for f in os.listdir(dir_path) if f.endswith('.txt')]
    if not txt_files:
        print(f"Error: No .txt files found in {dir_path}")
        sys.exit(1)

    # Use the first .txt file found (usually _chat.txt)
    txt_file = os.path.join(dir_path, txt_files[0])
    
    # Output name is the directory name
    dir_name = os.path.basename(dir_path)
    if not dir_name: # handles trailing slash
        dir_name = os.path.basename(os.path.dirname(dir_path))
    
    output_xlsx = f"{dir_name}.xlsx"
    output_path = os.path.join(os.path.dirname(dir_path), output_xlsx)

    print(f"Parsing {txt_file}...")
    messages = parse_chat(txt_file)
    create_xlsx(messages, output_path, sheet_name=dir_name)

if __name__ == '__main__':
    main()
